import json
import os
import shutil
import zipfile
from datetime import datetime
import xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

BODY_TOOLTIPS = {
    "Sedan": "Классический закрытый кузов с 4 дверями и отдельным багажником.",
    "Hatchback": "Короткий кузов с задней подъемной дверью и общим салоном с багажником.",
    "Wagon": "Удлиненный кузов с увеличенным багажным отсеком, как у универсала.",
    "Pickup": "Грузовая платформа сзади и отдельная кабина спереди, как у пикапа.",
    "Coupe": "Обычно 2-дверный кузов со спортивным силуэтом.",
    "Convertible": "Открытый кузов со складной или съемной крышей.",
    "Fastback": "Крыша плавно уходит назад одной линией до хвоста машины.",
    "Shooting Brake": "Спортивное купе с удлиненной задней частью, смесь купе и универсала.",
    "Liftback": "Похож на седан, но задняя дверь поднимается вместе со стеклом.",
    "Formula 1": "Одноместный открытый кузов болида с вынесенными колесами."
}

TYPE_TOOLTIPS = {
    "City Car": "Компактная городская машина для повседневной езды.",
    "Sports Car": "Легковая спортивная машина с акцентом на управляемость и скорость.",
    "Muscle Car": "Мощное американское купе или седан с большим мотором.",
    "Supercar": "Очень быстрый и дорогой спортивный автомобиль высокого класса.",
    "Hypercar": "Экстремально быстрый и технологичный автомобиль выше уровня supercar.",
    "Open Wheel": "Гоночная машина с открытыми колесами, как у формульных серий.",
    "GT": "Гоночный или дорожный grand tourer для скорости на длинных дистанциях.",
    "Rally": "Машина, подготовленная для раллийных трасс и смешанных покрытий.",
    "Touring Car": "Силуэтная гоночная машина, близкая по форме к серийным авто.",
    "Prototype": "Специально построенный гоночный автомобиль, не основанный на серийной модели."
}

SPECIAL_TOOLTIPS = {
    "Track-only": "Версия, созданная в первую очередь для трека и максимальной скорости на трассе.",
    "Tuned": "Модифицированный автомобиль с доработками внешности, шасси или двигателя.",
    "Urban Outlaw": "Харизматичный дорожный проект в стиле restomod или кастомной городской легенды.",
    "Classic": "Классическая модель, ценимая за возраст, стиль и исторический характер.",
    "Retro": "Автомобиль с ярко выраженной стилистикой прошлых десятилетий.",
    "Gulf Livery": "Окраска в знаменитой гоночной ливрее Gulf.",
    "Falken Livery": "Окраска в фирменной гоночной ливрее Falken.",
    "Borla Livery": "Окраска в стиле гоночной ливреи Borla.",
    "Advan Racing Livery": "Окраска в фирменном гоночном стиле Advan Racing.",
    "Streat Racing": "Автомобиль в духе уличных гонок, с акцентом на скорость, стиль и дорожный характер.",
    "Police": "Полицейская версия или стилизация под служебный автомобиль."
}

class CarDatabase:
    EXPORT_COLUMNS = [
        ("name", "Name"),
        ("make", "Make"),
        ("model", "Model"),
        ("color", "Color"),
        ("weight_g", "Weight (g)"),
        ("sku", "SKU"),
        ("brand", "Brand"),
        ("Body", "Body"),
        ("Type", "Type"),
        ("Special", "Special")
    ]

    def __init__(self, json_file="cars.json"):
        self.json_file = json_file
        self.cars_data = self.load_cars_database()
        self.reference_options = self.load_reference_options()
        self.ensure_reference_metadata_persisted()
        self.cars_data = self.load_cars_database()
        self.reference_options = self.load_reference_options()

    def _get_default_reference_options(self):
        return {
            "Body": [],
            "Type": [],
            "Special": [],
            "Make": [],
            "Brand": [],
            "BodyDescriptions": dict(BODY_TOOLTIPS),
            "TypeDescriptions": dict(TYPE_TOOLTIPS),
            "SpecialDescriptions": dict(SPECIAL_TOOLTIPS)
        }

    def _normalize_description_map(self, values):
        if not isinstance(values, dict):
            return {}

        normalized = {}
        for key, value in values.items():
            key_text = str(key).strip()
            value_text = str(value).strip()
            if key_text:
                normalized[key_text] = value_text
        return normalized

    def _find_image_path_by_sku(self, sku):
        sku_text = str(sku).strip()
        if not sku_text:
            return ""

        default_path = os.path.join("car_images", f"{sku_text}.webp").replace("\\", "/")
        if not os.path.isdir("car_images"):
            return default_path

        preferred_extensions = [".webp", ".png", ".jpg", ".jpeg"]
        matched_files = []

        try:
            for entry in os.scandir("car_images"):
                if not entry.is_file():
                    continue

                stem, extension = os.path.splitext(entry.name)
                if stem.lower() == sku_text.lower():
                    matched_files.append((entry.name, extension.lower()))
        except Exception:
            return default_path

        if not matched_files:
            return default_path

        matched_files.sort(
            key=lambda item: (
                preferred_extensions.index(item[1]) if item[1] in preferred_extensions else len(preferred_extensions),
                item[0].lower()
            )
        )
        return os.path.join("car_images", matched_files[0][0]).replace("\\", "/")

    def _normalize_multi_value_field(self, value):
        if isinstance(value, list):
            raw_values = value
        elif value is None:
            raw_values = []
        else:
            raw_values = str(value).replace(",", ";").split(";")

        normalized = []
        seen_values = set()
        for item in raw_values:
            text = str(item).strip()
            lowered_text = text.lower()
            if not text or lowered_text in seen_values:
                continue
            seen_values.add(lowered_text)
            normalized.append(text)
        return normalized

    def _normalize_imported_car(self, raw_car):
        if not isinstance(raw_car, dict):
            return None

        normalized_name = str(raw_car.get("name", "")).strip()
        normalized_sku = str(raw_car.get("sku", "")).strip()
        existing_image = ""

        for existing_car in self.cars_data:
            if not isinstance(existing_car, dict):
                continue

            existing_sku = str(existing_car.get("sku", "")).strip()
            existing_name = str(existing_car.get("name", "")).strip()

            if normalized_sku and existing_sku.lower() == normalized_sku.lower():
                existing_image = str(existing_car.get("image", "")).strip()
                break

            if normalized_name and existing_name.lower() == normalized_name.lower():
                existing_image = str(existing_car.get("image", "")).strip()
                break

        normalized_car = {
            "name": normalized_name,
            "make": str(raw_car.get("make", "")).strip(),
            "model": str(raw_car.get("model", "")).strip(),
            "color": str(raw_car.get("color", "")).strip(),
            "brand": str(raw_car.get("brand", "")).strip(),
            "sku": normalized_sku,
            "image": str(raw_car.get("image", existing_image)).strip(),
            "Body": self._normalize_multi_value_field(raw_car.get("Body", [])),
            "Type": self._normalize_multi_value_field(raw_car.get("Type", [])),
            "Special": self._normalize_multi_value_field(raw_car.get("Special", []))
        }

        if not normalized_car["image"] and normalized_sku:
            normalized_car["image"] = self._find_image_path_by_sku(normalized_sku)

        weight_value = raw_car.get("weight_g", "")
        if weight_value in ("", None):
            normalized_car["weight_g"] = ""
        else:
            try:
                numeric_weight = float(str(weight_value).replace(",", ".").strip())
                normalized_car["weight_g"] = int(numeric_weight) if numeric_weight.is_integer() else numeric_weight
            except Exception:
                normalized_car["weight_g"] = str(weight_value).strip()

        return normalized_car

    def backfill_missing_image_paths(self):
        try:
            with open(self.json_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, list):
                return 0

            updated_count = 0
            for item in data:
                if not isinstance(item, dict) or item.get("_meta") == "REFERENCE_OPTIONS":
                    continue

                current_image = str(item.get("image", "")).strip()
                if current_image:
                    continue

                sku = str(item.get("sku", "")).strip()
                inferred_image = self._find_image_path_by_sku(sku)
                if not inferred_image:
                    continue

                item["image"] = inferred_image
                updated_count += 1

            if not updated_count:
                return 0

            with open(self.json_file, "w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=2)

            self.reload_data()
            return updated_count

        except Exception as e:
            print("Ошибка автопривязки изображений: " + str(e))
            return 0

    def _build_reference_options_from_cars(self, cars):
        updated_reference_options = dict(self.reference_options)

        for key, field_name in (("Body", "Body"), ("Type", "Type"), ("Special", "Special")):
            existing_values = updated_reference_options.get(key, [])
            merged_values = {
                str(value).strip()
                for value in existing_values
                if str(value).strip()
            }

            for car in cars:
                for value in car.get(field_name, []):
                    text = str(value).strip()
                    if text:
                        merged_values.add(text)

            updated_reference_options[key] = sorted(merged_values, key=lambda value: value.lower())

        existing_brands = updated_reference_options.get("Brand", [])
        merged_brands = {
            str(value).strip()
            for value in existing_brands
            if str(value).strip()
        }

        for car in cars:
            brand = str(car.get("brand", "")).strip()
            if brand:
                merged_brands.add(brand)

        updated_reference_options["Brand"] = sorted(merged_brands, key=lambda value: value.lower())
        return updated_reference_options

    def _validate_imported_cars(self, cars):
        if not cars:
            return False, "В таблице не найдено ни одной машинки."

        name_map = {}
        sku_map = {}
        for index, car in enumerate(cars, start=2):
            name = str(car.get("name", "")).strip()
            sku = str(car.get("sku", "")).strip()

            if not name:
                return False, f"Строка {index}: поле Name не заполнено."

            lowered_name = name.lower()
            if lowered_name in name_map:
                return False, f'Дубликат Name в строках {name_map[lowered_name]} и {index}: "{name}".'
            name_map[lowered_name] = index

            if sku:
                lowered_sku = sku.lower()
                if lowered_sku in sku_map:
                    return False, f'Dубликат SKU в строках {sku_map[lowered_sku]} и {index}: "{sku}".'
                sku_map[lowered_sku] = index

        return True, ""

    def _create_backup_file(self):
        base_name = os.path.splitext(os.path.basename(self.json_file))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_path = f"{base_name}_backup_{timestamp}.json"
        shutil.copyfile(self.json_file, backup_path)
        return backup_path

    def _serialize_car_rows(self):
        rows = []
        header_row = [title for _, title in self.EXPORT_COLUMNS]
        rows.append(header_row)

        for car in self.cars_data:
            row = []
            for field_name, _ in self.EXPORT_COLUMNS:
                raw_value = car.get(field_name, "")
                if isinstance(raw_value, list):
                    value = "; ".join(str(item).strip() for item in raw_value if str(item).strip())
                else:
                    value = "" if raw_value is None else str(raw_value).strip()
                row.append(value)
            rows.append(row)

        return rows

    def _serialize_reference_rows(self):
        rows = [[
            "Body",
            "Body Description",
            "Type",
            "Type Description",
            "Special",
            "Special Description",
            "Brand"
        ]]

        body_values = list(self.reference_options.get("Body", []))
        type_values = list(self.reference_options.get("Type", []))
        special_values = list(self.reference_options.get("Special", []))
        brand_values = list(self.reference_options.get("Brand", []))
        body_descriptions = self.reference_options.get("BodyDescriptions", {})
        type_descriptions = self.reference_options.get("TypeDescriptions", {})
        special_descriptions = self.reference_options.get("SpecialDescriptions", {})

        max_length = max(len(body_values), len(type_values), len(special_values), len(brand_values), 1)
        for index in range(max_length):
            body_value = body_values[index] if index < len(body_values) else ""
            type_value = type_values[index] if index < len(type_values) else ""
            special_value = special_values[index] if index < len(special_values) else ""
            brand_value = brand_values[index] if index < len(brand_values) else ""

            rows.append([
                body_value,
                body_descriptions.get(body_value, "") if body_value else "",
                type_value,
                type_descriptions.get(type_value, "") if type_value else "",
                special_value,
                special_descriptions.get(special_value, "") if special_value else "",
                brand_value
            ])

        return rows

    def _write_imported_cars(self, imported_cars):
        is_valid, validation_message = self._validate_imported_cars(imported_cars)
        if not is_valid:
            return False, validation_message

        merged_cars = []
        existing_cars = [
            dict(car)
            for car in self.cars_data
            if isinstance(car, dict) and car.get("name")
        ]

        existing_by_sku = {}
        for index, car in enumerate(existing_cars):
            sku = str(car.get("sku", "")).strip().lower()
            if sku and sku not in existing_by_sku:
                existing_by_sku[sku] = index

        updated_count = 0
        added_count = 0

        for imported_car in imported_cars:
            sku = str(imported_car.get("sku", "")).strip().lower()

            target_index = None
            if sku and sku in existing_by_sku:
                target_index = existing_by_sku[sku]

            if target_index is None:
                existing_cars.append(imported_car)
                new_index = len(existing_cars) - 1
                if sku:
                    existing_by_sku[sku] = new_index
                added_count += 1
            else:
                existing_cars[target_index] = imported_car
                if sku:
                    existing_by_sku[sku] = target_index
                updated_count += 1

        merged_cars.extend(existing_cars)

        updated_reference_options = self._build_reference_options_from_cars(merged_cars)
        backup_path = self._create_backup_file()

        new_data = [{
            "_meta": "REFERENCE_OPTIONS",
            "Body": updated_reference_options.get("Body", []),
            "Type": updated_reference_options.get("Type", []),
            "Special": updated_reference_options.get("Special", []),
            "Brand": updated_reference_options.get("Brand", []),
            "BodyDescriptions": self._normalize_description_map(updated_reference_options.get("BodyDescriptions", {})),
            "TypeDescriptions": self._normalize_description_map(updated_reference_options.get("TypeDescriptions", {})),
            "SpecialDescriptions": self._normalize_description_map(updated_reference_options.get("SpecialDescriptions", {}))
        }]
        new_data.extend(merged_cars)

        with open(self.json_file, "w", encoding="utf-8") as file:
            json.dump(new_data, file, ensure_ascii=False, indent=2)

        self.reload_data()
        return True, (
            f"Импорт завершён. "
            f"Добавлено: {added_count}, обновлено: {updated_count}, "
            f"всего в файле: {len(imported_cars)}. "
            f"Резервная копия: {backup_path}"
        )

    def _column_letter(self, column_number):
        result = ""
        current = column_number
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _excel_escape(self, value):
        return (
            str(value)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )

    def _build_xlsx_sheet_xml(self, rows):
        max_columns = max((len(row) for row in rows), default=1)
        lines = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
            '  <sheetViews><sheetView workbookViewId="0"/></sheetViews>',
            '  <sheetFormatPr defaultRowHeight="15"/>',
            f'  <dimension ref="A1:{self._column_letter(max_columns)}{len(rows)}"/>',
            '  <sheetData>'
        ]

        for row_index, row_values in enumerate(rows, start=1):
            lines.append(f'    <row r="{row_index}">')
            for column_index, value in enumerate(row_values, start=1):
                cell_ref = f"{self._column_letter(column_index)}{row_index}"
                escaped_value = self._excel_escape(value)
                lines.append(
                    f'      <c r="{cell_ref}" t="inlineStr"><is><t>{escaped_value}</t></is></c>'
                )
            lines.append('    </row>')

        lines.extend([
            '  </sheetData>',
            '</worksheet>'
        ])
        return "\n".join(lines)

    def _read_xlsx_rows(self, file_path):
        ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
            "docrel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        }

        with zipfile.ZipFile(file_path, "r") as archive:
            workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
            workbook_rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

            relationship_map = {}
            for rel in workbook_rels_xml.findall("rel:Relationship", ns):
                relationship_map[rel.get("Id")] = rel.get("Target")

            sheet_path = None
            for sheet in workbook_xml.findall("main:sheets/main:sheet", ns):
                if sheet.get("name") == "Cars":
                    relationship_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                    target = relationship_map.get(relationship_id, "")
                    if target:
                        sheet_path = "xl/" + target.lstrip("/")
                    break

            if sheet_path is None:
                first_sheet = workbook_xml.find("main:sheets/main:sheet", ns)
                if first_sheet is None:
                    raise ValueError("В XLSX-файле не найдено ни одного листа.")
                relationship_id = first_sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                target = relationship_map.get(relationship_id, "")
                if not target:
                    raise ValueError("Не удалось определить путь к листу в XLSX-файле.")
                sheet_path = "xl/" + target.lstrip("/")

            shared_strings = []
            if "xl/sharedStrings.xml" in archive.namelist():
                shared_xml = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                for si in shared_xml.findall("main:si", ns):
                    text_parts = []
                    plain_text = si.find("main:t", ns)
                    if plain_text is not None and plain_text.text is not None:
                        text_parts.append(plain_text.text)
                    for run_text in si.findall("main:r/main:t", ns):
                        if run_text.text is not None:
                            text_parts.append(run_text.text)
                    shared_strings.append("".join(text_parts))

            sheet_xml = ET.fromstring(archive.read(sheet_path))
            rows = []

            for row in sheet_xml.findall(".//main:sheetData/main:row", ns):
                row_values = {}
                max_column = 0

                for cell in row.findall("main:c", ns):
                    cell_ref = cell.get("r", "")
                    column_letters = "".join(ch for ch in cell_ref if ch.isalpha())
                    if not column_letters:
                        continue

                    column_number = 0
                    for letter in column_letters:
                        column_number = (column_number * 26) + (ord(letter.upper()) - 64)
                    max_column = max(max_column, column_number)

                    cell_type = cell.get("t", "")
                    value = ""

                    if cell_type == "inlineStr":
                        inline_text = cell.find("main:is/main:t", ns)
                        if inline_text is not None and inline_text.text is not None:
                            value = inline_text.text
                    elif cell_type == "s":
                        value_node = cell.find("main:v", ns)
                        if value_node is not None and value_node.text is not None:
                            shared_index = int(value_node.text)
                            if 0 <= shared_index < len(shared_strings):
                                value = shared_strings[shared_index]
                    else:
                        value_node = cell.find("main:v", ns)
                        if value_node is not None and value_node.text is not None:
                            value = value_node.text

                    row_values[column_number] = value.strip()

                if max_column == 0:
                    continue

                rows.append([row_values.get(index, "") for index in range(1, max_column + 1)])

        return rows

    def _parse_rows_to_cars(self, rows):
        if len(rows) < 2:
            return False, "В Excel-файле нет строк с данными."

        headers = [str(value).strip() for value in rows[0]]
        header_map = {header.lower(): index for index, header in enumerate(headers) if header}
        required_headers = [title for _, title in self.EXPORT_COLUMNS]
        missing_headers = [header for header in required_headers if header.lower() not in header_map]
        if missing_headers:
            return False, "Не найдены столбцы: " + ", ".join(missing_headers)

        imported_cars = []
        for row in rows[1:]:
            raw_car = {}
            for field_name, title in self.EXPORT_COLUMNS:
                column_index = header_map[title.lower()]
                raw_car[field_name] = row[column_index] if column_index < len(row) else ""

            normalized_car = self._normalize_imported_car(raw_car)
            if normalized_car and any(str(value).strip() for value in normalized_car.values() if not isinstance(value, list)):
                imported_cars.append(normalized_car)

        return True, imported_cars

    def export_to_xlsx_file(self, file_path):
        if Workbook is not None:
            try:
                workbook = Workbook()
                cars_sheet = workbook.active
                cars_sheet.title = "Cars"

                for row_values in self._serialize_car_rows():
                    cars_sheet.append(row_values)

                reference_sheet = workbook.create_sheet("Reference")
                for row_values in self._serialize_reference_rows():
                    reference_sheet.append(row_values)

                workbook.save(file_path)
                return True, f"Экспортировано машинок: {len(self.cars_data)}"

            except Exception as e:
                return False, f"Ошибка экспорта XLSX: {e}"

        try:
            car_rows = self._serialize_car_rows()
            reference_rows = self._serialize_reference_rows()
            cars_sheet_xml = self._build_xlsx_sheet_xml(car_rows)
            reference_sheet_xml = self._build_xlsx_sheet_xml(reference_rows)

            content_types_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>"""

            rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>"""

            workbook_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Cars" sheetId="1" r:id="rId1"/>
    <sheet name="Reference" sheetId="2" r:id="rId2"/>
  </sheets>
</workbook>"""

            workbook_rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>
</Relationships>"""

            created_iso = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
            core_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
 xmlns:dc="http://purl.org/dc/elements/1.1/"
 xmlns:dcterms="http://purl.org/dc/terms/"
 xmlns:dcmitype="http://purl.org/dc/dcmitype/"
 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>Hot Wheels Timer</dc:creator>
  <cp:lastModifiedBy>Hot Wheels Timer</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">{created_iso}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{created_iso}</dcterms:modified>
</cp:coreProperties>"""

            app_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
 xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Hot Wheels Timer</Application>
  <TitlesOfParts>
    <vt:vector size="2" baseType="lpstr">
      <vt:lpstr>Cars</vt:lpstr>
      <vt:lpstr>Reference</vt:lpstr>
    </vt:vector>
  </TitlesOfParts>
  <HeadingPairs>
    <vt:vector size="2" baseType="variant">
      <vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant>
      <vt:variant><vt:i4>2</vt:i4></vt:variant>
    </vt:vector>
  </HeadingPairs>
</Properties>"""

            with zipfile.ZipFile(file_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("[Content_Types].xml", content_types_xml)
                archive.writestr("_rels/.rels", rels_xml)
                archive.writestr("xl/workbook.xml", workbook_xml)
                archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
                archive.writestr("xl/worksheets/sheet1.xml", cars_sheet_xml)
                archive.writestr("xl/worksheets/sheet2.xml", reference_sheet_xml)
                archive.writestr("docProps/core.xml", core_xml)
                archive.writestr("docProps/app.xml", app_xml)

            return True, f"Экспортировано машинок: {len(self.cars_data)}"

        except Exception as e:
            return False, f"Ошибка экспорта XLSX: {e}"

    def import_from_xlsx_file(self, file_path):
        if load_workbook is not None:
            workbook = None
            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                sheet = workbook["Cars"] if "Cars" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(["" if value is None else str(value).strip() for value in row])
                success, parsed = self._parse_rows_to_cars(rows)
                if not success:
                    return False, parsed
                return self._write_imported_cars(parsed)
            except Exception as e:
                return False, f"Ошибка импорта XLSX: {e}"
            finally:
                if workbook is not None:
                    workbook.close()

        try:
            rows = self._read_xlsx_rows(file_path)
            success, parsed = self._parse_rows_to_cars(rows)
            if not success:
                return False, parsed
            return self._write_imported_cars(parsed)
        except zipfile.BadZipFile:
            return False, "Не удалось прочитать XLSX-файл. Файл повреждён или имеет неверный формат."
        except Exception as e:
            return False, f"Ошибка импорта XLSX: {e}"

    def export_to_excel_file(self, file_path):
        try:
            if str(file_path).lower().endswith(".xlsx"):
                return self.export_to_xlsx_file(file_path)

            ET.register_namespace("", "urn:schemas-microsoft-com:office:spreadsheet")
            ET.register_namespace("o", "urn:schemas-microsoft-com:office:office")
            ET.register_namespace("x", "urn:schemas-microsoft-com:office:excel")
            ET.register_namespace("ss", "urn:schemas-microsoft-com:office:spreadsheet")

            ns = {
                "ss": "urn:schemas-microsoft-com:office:spreadsheet"
            }

            workbook = ET.Element(
                "{urn:schemas-microsoft-com:office:spreadsheet}Workbook",
                {
                    "{urn:schemas-microsoft-com:office:office}ProgId": "Excel.Sheet",
                    "{urn:schemas-microsoft-com:office:excel}FullColumns": "1",
                    "{urn:schemas-microsoft-com:office:excel}FullRows": "1"
                }
            )

            worksheet = ET.SubElement(
                workbook,
                "{urn:schemas-microsoft-com:office:spreadsheet}Worksheet",
                {f"{{{ns['ss']}}}Name": "Cars"}
            )
            table = ET.SubElement(worksheet, "{urn:schemas-microsoft-com:office:spreadsheet}Table")

            header_row = ET.SubElement(table, "{urn:schemas-microsoft-com:office:spreadsheet}Row")
            for _, title in self.EXPORT_COLUMNS:
                cell = ET.SubElement(header_row, "{urn:schemas-microsoft-com:office:spreadsheet}Cell")
                data = ET.SubElement(
                    cell,
                    "{urn:schemas-microsoft-com:office:spreadsheet}Data",
                    {f"{{{ns['ss']}}}Type": "String"}
                )
                data.text = title

            for car in self.cars_data:
                row = ET.SubElement(table, "{urn:schemas-microsoft-com:office:spreadsheet}Row")
                for field_name, _ in self.EXPORT_COLUMNS:
                    raw_value = car.get(field_name, "")
                    if isinstance(raw_value, list):
                        value = "; ".join(str(item).strip() for item in raw_value if str(item).strip())
                        value_type = "String"
                    elif field_name == "weight_g" and raw_value not in ("", None):
                        value = str(raw_value)
                        value_type = "Number"
                    else:
                        value = str(raw_value).strip()
                        value_type = "String"

                    cell = ET.SubElement(row, "{urn:schemas-microsoft-com:office:spreadsheet}Cell")
                    data = ET.SubElement(
                        cell,
                        "{urn:schemas-microsoft-com:office:spreadsheet}Data",
                        {f"{{{ns['ss']}}}Type": value_type}
                    )
                    data.text = value

            reference_worksheet = ET.SubElement(
                workbook,
                "{urn:schemas-microsoft-com:office:spreadsheet}Worksheet",
                {f"{{{ns['ss']}}}Name": "Reference"}
            )
            reference_table = ET.SubElement(
                reference_worksheet,
                "{urn:schemas-microsoft-com:office:spreadsheet}Table"
            )

            for row_values in self._serialize_reference_rows():
                row = ET.SubElement(reference_table, "{urn:schemas-microsoft-com:office:spreadsheet}Row")
                for value in row_values:
                    cell = ET.SubElement(row, "{urn:schemas-microsoft-com:office:spreadsheet}Cell")
                    data = ET.SubElement(
                        cell,
                        "{urn:schemas-microsoft-com:office:spreadsheet}Data",
                        {f"{{{ns['ss']}}}Type": "String"}
                    )
                    data.text = str(value).strip()

            tree = ET.ElementTree(workbook)
            with open(file_path, "wb") as file:
                file.write(b'<?xml version="1.0" encoding="utf-8"?>\n')
                tree.write(file, encoding="utf-8", xml_declaration=False)

            return True, f"Экспортировано машинок: {len(self.cars_data)}"

        except Exception as e:
            return False, f"Ошибка экспорта: {e}"

    def import_from_excel_file(self, file_path):
        try:
            if str(file_path).lower().endswith(".xlsx"):
                return self.import_from_xlsx_file(file_path)

            tree = ET.parse(file_path)
            root = tree.getroot()
            namespace = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
            rows = root.findall(".//ss:Worksheet[@ss:Name='Cars']/ss:Table/ss:Row", namespace)
            if not rows:
                rows = root.findall(".//ss:Table/ss:Row", namespace)

            parsed_rows = []
            for row in rows:
                cells = row.findall("ss:Cell", namespace)
                values = []
                for cell in cells:
                    data = cell.find("ss:Data", namespace)
                    values.append("" if data is None or data.text is None else str(data.text).strip())
                parsed_rows.append(values)

            success, parsed = self._parse_rows_to_cars(parsed_rows)
            if not success:
                return False, parsed
            return self._write_imported_cars(parsed)

        except ET.ParseError:
            return False, "Не удалось прочитать Excel XML-файл. Ожидается файл, экспортированный из программы."
        except Exception as e:
            return False, f"Ошибка импорта: {e}"

    def load_cars_database(self):
        try:
            with open(self.json_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, list):
                return []

            cars_list = []

            for item in data:
                # Обычная запись-словарь
                if isinstance(item, dict):
                    # Пропускаем служебный блок со справочниками
                    if item.get("_meta") == "REFERENCE_OPTIONS":
                        continue

                    # Берём только реальные машины, у которых есть name
                    if item.get("name"):
                        cars_list.append(item)

                # Если по ошибке встретился вложенный список — тоже разбираем его
                elif isinstance(item, list):
                    for sub_item in item:
                        if isinstance(sub_item, dict) and sub_item.get("name"):
                            cars_list.append(sub_item)

            return cars_list

        except Exception as e:
            print("Ошибка загрузки cars.json:", e)
            return []

    def load_reference_options(self):
        try:
            with open(self.json_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, list):
                return self._get_default_reference_options()

            # Популярные мировые производители автомобилей
            popular_makes = [
                "Ferrari", "Lamborghini", "Porsche", "Mercedes-Benz", "BMW",
                "Audi", "Volkswagen", "Aston Martin", "Bentley", "Rolls-Royce",
                "Bugatti", "McLaren", "Pagani", "Koenigsegg", "Hennessey",
                "Lotus", "Jaguar", "Land Rover", "Maserati", "Alfa Romeo",
                "Fiat", "Lancia", "Bugatti", "Bugatti", "Citroën", "Renault",
                "Peugeot", "Toyota", "Honda", "Nissan", "Mazda", "Subaru",
                "Mitsubishi", "Hyundai", "Kia", "Tata", "Mahindra",
                "Ford", "General Motors", "Chrysler", "Tesla", "Rivian",
                "Lucid", "Polestar", "Volvo", "Saab", "Geely", "BYD",
                "GAC Aion", "NIO", "XPeng", "Li Auto", "Changan",
                "Chery", "Great Wall", "Haval", "SAIC", "FAW"
            ]

            # Собираем существующие производители из базы данных
            existing_makes = set()
            existing_brands = set()
            for car in self.cars_data:
                if isinstance(car, dict):
                    make_value = str(car.get("make", "")).strip()
                    brand_value = str(car.get("brand", "")).strip()
                    if make_value:
                        existing_makes.add(make_value)
                    if brand_value:
                        existing_brands.add(brand_value)

            # Объединяем: существующие + популярные, убираем дубликаты
            all_makes = sorted(set(list(existing_makes) + popular_makes), key=lambda x: x.lower())
            all_brands = sorted(existing_brands, key=lambda x: x.lower())

            body_list = []
            type_list = []
            special_list = []
            brand_list = []
            body_descriptions = dict(BODY_TOOLTIPS)
            type_descriptions = dict(TYPE_TOOLTIPS)
            special_descriptions = dict(SPECIAL_TOOLTIPS)

            for item in data:
                if isinstance(item, dict) and item.get("_meta") == "REFERENCE_OPTIONS":
                    body_list = item.get("Body", [])
                    type_list = item.get("Type", [])
                    special_list = item.get("Special", [])
                    brand_list = item.get("Brand", [])
                    body_descriptions.update(self._normalize_description_map(item.get("BodyDescriptions", {})))
                    type_descriptions.update(self._normalize_description_map(item.get("TypeDescriptions", {})))
                    special_descriptions.update(self._normalize_description_map(item.get("SpecialDescriptions", {})))

                    # На всякий случай приводим к спискам
                    if not isinstance(body_list, list):
                        body_list = []
                    if not isinstance(type_list, list):
                        type_list = []
                    if not isinstance(special_list, list):
                        special_list = []
                    if not isinstance(brand_list, list):
                        brand_list = []

                    break

            if not brand_list:
                brand_list = all_brands

            return {
                "Body": body_list,
                "Type": type_list,
                "Special": special_list,
                "Make": all_makes,
                "Brand": brand_list,
                "BodyDescriptions": body_descriptions,
                "TypeDescriptions": type_descriptions,
                "SpecialDescriptions": special_descriptions
            }

        except Exception:
            return self._get_default_reference_options()

    def find_car_in_database(self, car_name):
        search_name = car_name.strip().lower()
        for car in self.cars_data:
            db_name = str(car.get("name", "")).strip().lower()
            if db_name == search_name:
                return car
        return None

    def get_car_extra_data(self, car_name):
        car_data = self.find_car_in_database(car_name)

        if car_data:
            sku = str(car_data.get("sku", "")).strip()
            brand = str(car_data.get("brand", "")).strip()

            return {
                "sku": sku,
                "brand": brand
            }

        return {
            "sku": "",
            "brand": ""
        }

    def save_new_car_to_database(self, new_car_data):
        try:
            with open(self.json_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, list):
                return False

            new_name = str(new_car_data.get("name", "")).strip().lower()
            new_sku = str(new_car_data.get("sku", "")).strip().lower()

            for item in data:
                if not isinstance(item, dict):
                    continue

                existing_name = str(item.get("name", "")).strip().lower()
                existing_sku = str(item.get("sku", "")).strip().lower()

                if existing_name and existing_name == new_name:
                    return False  # Duplicate name

                if existing_sku and existing_sku == new_sku:
                    return False  # Duplicate SKU

            data.append(new_car_data)

            with open(self.json_file, "w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=2)

            return True

        except Exception as e:
            print("Ошибка записи новой машинки в cars.json: " + str(e))
            return False

    def save_edited_car_to_database(self, original_name, updated_car_data):
        try:
            with open(self.json_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, list):
                return False

            original_name_lower = str(original_name).strip().lower()
            updated_name_lower = str(updated_car_data.get("name", "")).strip().lower()
            updated_sku_lower = str(updated_car_data.get("sku", "")).strip().lower()

            target_index = -1

            for index, item in enumerate(data):
                if not isinstance(item, dict):
                    continue

                item_name = str(item.get("name", "")).strip().lower()
                item_sku = str(item.get("sku", "")).strip().lower()

                # Находим редактируемую запись
                if item_name == original_name_lower and target_index == -1:
                    target_index = index
                    continue

                # Проверка дубликата имени
                if item_name and item_name == updated_name_lower:
                    return False

                # Проверка дубликата SKU
                if item_sku and item_sku == updated_sku_lower:
                    return False

            if target_index < 0:
                return False

            data[target_index] = updated_car_data

            with open(self.json_file, "w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=2)

            return True

        except Exception as e:
            print("Ошибка сохранения изменений машинки: " + str(e))
            return False

    def save_reference_options(self, updated_reference_options, rename_operations=None):
        try:
            with open(self.json_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, list):
                return False

            normalized_options = {}
            for key in ("Body", "Type", "Special", "Brand"):
                raw_values = updated_reference_options.get(key, [])
                if not isinstance(raw_values, list):
                    raw_values = []

                cleaned_values = []
                seen_values = set()
                for value in raw_values:
                    text = str(value).strip()
                    lowered_text = text.lower()
                    if not text or lowered_text in seen_values:
                        continue
                    seen_values.add(lowered_text)
                    cleaned_values.append(text)

                cleaned_values.sort(key=lambda value: value.lower())
                normalized_options[key] = cleaned_values

            reference_block = {
                "_meta": "REFERENCE_OPTIONS",
                "Body": normalized_options["Body"],
                "Type": normalized_options["Type"],
                "Special": normalized_options["Special"],
                "Brand": normalized_options["Brand"],
                "BodyDescriptions": self._normalize_description_map(updated_reference_options.get("BodyDescriptions", {})),
                "TypeDescriptions": self._normalize_description_map(updated_reference_options.get("TypeDescriptions", {})),
                "SpecialDescriptions": self._normalize_description_map(updated_reference_options.get("SpecialDescriptions", {}))
            }

            reference_index = -1
            for index, item in enumerate(data):
                if isinstance(item, dict) and item.get("_meta") == "REFERENCE_OPTIONS":
                    reference_index = index
                    break

            rename_operations = rename_operations or {}
            for field_name, operations in rename_operations.items():
                if not isinstance(operations, list):
                    continue

                for operation in operations:
                    if not isinstance(operation, (list, tuple)) or len(operation) != 2:
                        continue

                    old_value = str(operation[0]).strip()
                    new_value = str(operation[1]).strip()
                    if not old_value or not new_value:
                        continue

                    for item in data:
                        if not isinstance(item, dict) or item.get("_meta") == "REFERENCE_OPTIONS":
                            continue

                        if field_name == "brand":
                            current_brand = str(item.get("brand", "")).strip()
                            if current_brand.lower() == old_value.lower():
                                item["brand"] = new_value
                            continue

                        field_values = item.get(field_name, [])
                        if not isinstance(field_values, list):
                            field_values = [field_values]

                        updated_values = []
                        changed = False
                        seen_values = set()

                        for field_value in field_values:
                            field_text = str(field_value).strip()
                            if field_text.lower() == old_value.lower():
                                field_text = new_value
                                changed = True

                            lowered_text = field_text.lower()
                            if field_text and lowered_text not in seen_values:
                                seen_values.add(lowered_text)
                                updated_values.append(field_text)

                        if changed:
                            item[field_name] = updated_values

            if reference_index >= 0:
                data[reference_index] = reference_block
            else:
                data.insert(0, reference_block)

            with open(self.json_file, "w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=2)

            self.reload_data()
            return True

        except Exception as e:
            print("Ошибка сохранения справочников: " + str(e))
            return False

    def ensure_reference_metadata_persisted(self):
        try:
            with open(self.json_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, list):
                return False

            reference_block = None
            for item in data:
                if isinstance(item, dict) and item.get("_meta") == "REFERENCE_OPTIONS":
                    reference_block = item
                    break

            if reference_block is None:
                return False

            changed = False

            if "Brand" not in reference_block:
                reference_block["Brand"] = list(self.reference_options.get("Brand", []))
                changed = True

            if "BodyDescriptions" not in reference_block:
                reference_block["BodyDescriptions"] = dict(self.reference_options.get("BodyDescriptions", {}))
                changed = True

            if "TypeDescriptions" not in reference_block:
                reference_block["TypeDescriptions"] = dict(self.reference_options.get("TypeDescriptions", {}))
                changed = True

            if "SpecialDescriptions" not in reference_block:
                reference_block["SpecialDescriptions"] = dict(self.reference_options.get("SpecialDescriptions", {}))
                changed = True

            if not changed:
                return False

            with open(self.json_file, "w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=2)

            return True

        except Exception as e:
            print("Ошибка миграции справочников: " + str(e))
            return False

    def reload_data(self):
        self.cars_data = self.load_cars_database()
        self.reference_options = self.load_reference_options()
